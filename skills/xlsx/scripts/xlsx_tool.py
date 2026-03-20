import csv
import json
import sys
from pathlib import Path


def _err(message: str, **extra):
    payload = {"success": False, "error": message}
    payload.update(extra)
    print(json.dumps(payload, ensure_ascii=False))


def _ok(**data):
    payload = {"success": True}
    payload.update(data)
    print(json.dumps(payload, ensure_ascii=False))


def _load_json_arg() -> dict:
    if len(sys.argv) < 2:
        return {}
    raw = sys.argv[1]
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)


def _read_workbook(input_path: str, data_only: bool = False, read_only: bool = False):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        _err("missing_dependency", detail=str(e))
        return None
    p = Path(input_path)
    if not p.is_file():
        _err("input_not_found", input=input_path)
        return None
    try:
        return load_workbook(filename=str(p), data_only=bool(data_only), read_only=bool(read_only))
    except Exception as e:
        _err("failed_to_read_xlsx", detail=str(e), input=str(p))
        return None


def _action_info(input_path: str) -> None:
    wb = _read_workbook(input_path, data_only=False, read_only=True)
    if wb is None:
        return
    try:
        sheets = list(wb.sheetnames or [])
        active = wb.active.title if getattr(wb, "active", None) else ""
        _ok(action="info", input=str(Path(input_path)), sheets=sheets, active=active)
    except Exception as e:
        _err("failed_to_inspect_xlsx", detail=str(e), input=input_path)


def _action_to_csv(input_path: str, output_path: str, sheet: str | None) -> None:
    wb = _read_workbook(input_path, data_only=True, read_only=True)
    if wb is None:
        return
    out = Path(output_path)
    _ensure_parent(out)
    try:
        ws = wb[sheet] if sheet else wb.active
    except Exception:
        _err("sheet_not_found", input=input_path, sheet=sheet or "")
        return
    try:
        with out.open("w", encoding="utf-8", errors="replace", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        _ok(action="to_csv", input=input_path, sheet=ws.title, output=str(out))
    except Exception as e:
        _err("failed_to_write_csv", detail=str(e), output=str(out))


def _action_from_csv(input_path: str, output_path: str, sheet: str) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        _err("missing_dependency", detail=str(e))
        return
    inp = Path(input_path)
    if not inp.is_file():
        _err("input_not_found", input=input_path)
        return
    out = Path(output_path)
    _ensure_parent(out)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet or "Sheet1"
        with inp.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        wb.save(str(out))
        _ok(action="from_csv", input=str(inp), output=str(out), sheet=ws.title)
    except Exception as e:
        _err("failed_to_create_xlsx", detail=str(e), input=str(inp), output=str(out))


def _action_preview_csv(input_path: str, max_rows: int) -> None:
    inp = Path(input_path)
    if not inp.is_file():
        _err("input_not_found", input=input_path)
        return
    try:
        with inp.open("r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            delimiter = "\t" if "\t" in sample and "," not in sample else ","
            reader = csv.reader(f, delimiter=delimiter)
            header = next(reader, [])
            rows: list[list[str]] = []
            for _ in range(max(0, int(max_rows))):
                row = next(reader, None)
                if row is None:
                    break
                rows.append([str(x) for x in row])
        _ok(action="preview_csv", input=str(inp), delimiter=delimiter, header=header, rows=rows)
    except Exception as e:
        _err("failed_to_preview_csv", detail=str(e), input=str(inp))


def _action_preview_sheet(input_path: str, sheet: str | None, max_rows: int, max_cols: int, data_only: bool) -> None:
    wb = _read_workbook(input_path, data_only=bool(data_only), read_only=True)
    if wb is None:
        return
    try:
        ws = wb[sheet] if sheet else wb.active
    except Exception:
        _err("sheet_not_found", input=input_path, sheet=sheet or "")
        return
    rows: list[list[str]] = []
    max_r = max(1, min(int(max_rows), 200))
    max_c = max(1, min(int(max_cols), 200))
    try:
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > max_r:
                break
            out_row: list[str] = []
            for c_idx, v in enumerate(row, start=1):
                if c_idx > max_c:
                    break
                if v is None:
                    out_row.append("")
                else:
                    out_row.append(str(v))
            rows.append(out_row)
        _ok(
            action="preview_sheet",
            input=input_path,
            sheet=ws.title,
            data_only=bool(data_only),
            max_rows=max_r,
            max_cols=max_c,
            rows=rows,
        )
    except Exception as e:
        _err("failed_to_preview_sheet", detail=str(e), input=input_path, sheet=ws.title)


def _action_get_cell(input_path: str, sheet: str | None, cell: str, data_only: bool) -> None:
    wb = _read_workbook(input_path, data_only=bool(data_only), read_only=True)
    if wb is None:
        return
    try:
        ws = wb[sheet] if sheet else wb.active
    except Exception:
        _err("sheet_not_found", input=input_path, sheet=sheet or "")
        return
    coord = (cell or "").strip()
    if not coord:
        _err("cell_required")
        return
    try:
        v = ws[coord].value
        _ok(action="get_cell", input=input_path, sheet=ws.title, cell=coord, data_only=bool(data_only), value=v)
    except Exception as e:
        _err("failed_to_get_cell", detail=str(e), input=input_path, sheet=ws.title, cell=coord)


def _action_set_cell(input_path: str, output_path: str | None, sheet: str | None, cell: str, value) -> None:
    wb = _read_workbook(input_path, data_only=False, read_only=False)
    if wb is None:
        return
    try:
        ws = wb[sheet] if sheet else wb.active
    except Exception:
        _err("sheet_not_found", input=input_path, sheet=sheet or "")
        return
    coord = (cell or "").strip()
    if not coord:
        _err("cell_required")
        return
    out = Path(output_path) if (output_path and str(output_path).strip()) else Path(input_path)
    _ensure_parent(out)
    try:
        ws[coord].value = value
        wb.save(str(out))
        _ok(action="set_cell", input=input_path, output=str(out), sheet=ws.title, cell=coord)
    except Exception as e:
        _err("failed_to_set_cell", detail=str(e), input=input_path, output=str(out), sheet=ws.title, cell=coord)


def _action_recalc(input_path: str, timeout: int) -> None:
    try:
        from recalc import recalc as _recalc
    except Exception as e:
        _err("missing_dependency", detail=str(e))
        return
    p = Path(input_path)
    if not p.is_file():
        _err("input_not_found", input=input_path)
        return
    try:
        result = _recalc(str(p), int(timeout))
        if isinstance(result, dict):
            result_payload = result
        else:
            result_payload = {"result": str(result)}
        _ok(action="recalc", input=str(p), result=result_payload)
    except Exception as e:
        _err("failed_to_recalc", detail=str(e), input=str(p))

def main() -> None:
    payload = _load_json_arg()
    action = str(payload.get("action") or "").strip()
    if not action:
        _err("missing_action")
        return

    if action == "info":
        _action_info(str(payload.get("input") or ""))
        return

    if action == "to_csv":
        _action_to_csv(
            str(payload.get("input") or ""),
            str(payload.get("output") or ""),
            str(payload.get("sheet") or "").strip() or None,
        )
        return

    if action == "from_csv":
        _action_from_csv(
            str(payload.get("input") or ""),
            str(payload.get("output") or ""),
            str(payload.get("sheet") or "Sheet1"),
        )
        return

    if action == "preview_csv":
        _action_preview_csv(
            str(payload.get("input") or ""),
            int(payload.get("max_rows") or 20),
        )
        return

    if action == "preview_sheet":
        _action_preview_sheet(
            str(payload.get("input") or ""),
            str(payload.get("sheet") or "").strip() or None,
            int(payload.get("max_rows") or 20),
            int(payload.get("max_cols") or 30),
            bool(payload.get("data_only") if "data_only" in payload else True),
        )
        return

    if action == "get_cell":
        _action_get_cell(
            str(payload.get("input") or ""),
            str(payload.get("sheet") or "").strip() or None,
            str(payload.get("cell") or ""),
            bool(payload.get("data_only") if "data_only" in payload else True),
        )
        return

    if action == "set_cell":
        _action_set_cell(
            str(payload.get("input") or ""),
            str(payload.get("output") or "").strip() or None,
            str(payload.get("sheet") or "").strip() or None,
            str(payload.get("cell") or ""),
            payload.get("value"),
        )
        return

    if action == "recalc":
        _action_recalc(
            str(payload.get("input") or ""),
            int(payload.get("timeout") or 30),
        )
        return

    _err("unsupported_action", action=action)


if __name__ == "__main__":
    main()
